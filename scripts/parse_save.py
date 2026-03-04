import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from collections import defaultdict

# Parse the save file
# Run from the repo root:  python scripts/parse_save.py
import pathlib
_repo = pathlib.Path(__file__).parent.parent
_saves = _repo / "saves"
_save_folder = max(_saves.iterdir(), key=lambda f: f.stat().st_mtime)
save_file = str(_save_folder / "SaveGameInfo")
tree = ET.parse(save_file)
root = tree.getroot()

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "Save Data"

# Define header style
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")

# Add headers
ws['A1'] = "Attribute Name"
ws['B1'] = "Value"
ws['C1'] = "Type/Category"

for col in ['A1', 'B1', 'C1']:
    ws[col].fill = header_fill
    ws[col].font = header_font
    ws[col].alignment = Alignment(horizontal="center", vertical="center")

# Recursively extract all elements
attributes = []

def extract_elements(element, path=""):
    """Recursively extract all XML elements and their values"""
    current_path = f"{path}/{element.tag}" if path else element.tag
    
    # Add element with its text value
    value = element.text.strip() if element.text else ""
    if value or len(element) == 0:  # Include if has value or is leaf node
        attributes.append({
            'name': current_path,
            'value': value[:100] if value else "(empty)",  # Truncate long values
            'type': 'Element'
        })
    
    # Process attributes
    for attr_name, attr_value in element.attrib.items():
        attributes.append({
            'name': f"{current_path}@{attr_name}",
            'value': attr_value[:100],
            'type': 'Attribute'
        })
    
    # Recursively process children
    for child in element:
        extract_elements(child, current_path)

# Extract all elements
extract_elements(root)

# Sort by name for better organization
attributes.sort(key=lambda x: x['name'])

# Write to Excel
row = 2
for attr in attributes:
    ws[f'A{row}'] = attr['name']
    ws[f'B{row}'] = attr['value']
    ws[f'C{row}'] = attr['type']
    row += 1

# Adjust column widths
ws.column_dimensions['A'].width = 50
ws.column_dimensions['B'].width = 40
ws.column_dimensions['C'].width = 15

# Save workbook to output/
output_file = str(_repo / "output" / "Stardew_Save_Attributes.xlsx")
wb.save(output_file)

print(f"✓ Excel file created: {output_file}")
print(f"✓ Total attributes extracted: {len(attributes)}")
